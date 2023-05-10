import streamlit as st
from openpyxl import load_workbook

st.title("Sondage Ethic Fashion")

# Ouvrir le fichier Excel existant
wb = load_workbook('data_base.xlsx')

# Sélectionner la feuille de calcul
sheet = wb.active

def process_data(data1, data2, data3, data4, data5, data6, data7, data8, data9, data10):
    # Trouver la prochaine ligne vide pour insérer les données
    next_row = sheet.max_row + 1

    # Insérer les données dans le fichier Excel
    sheet.cell(row=next_row, column=1, value=data1)
    sheet.cell(row=next_row, column=2, value=data2)
    sheet.cell(row=next_row, column=3, value=data3)
    sheet.cell(row=next_row, column=4, value=data4)
    sheet.cell(row=next_row, column=5, value=data5)
    sheet.cell(row=next_row, column=6, value=data6)
    sheet.cell(row=next_row, column=7, value=data7)
    sheet.cell(row=next_row, column=8, value=data8)
    sheet.cell(row=next_row, column=9, value=data9)
    sheet.cell(row=next_row, column=10, value=data10)

    # Enregistrer le fichier Excel
    wb.save('data_base.xlsx')


def main():
    # Afficher le formulaire 
    data1  = st.text_input("Variable 1")
    data2  = st.text_input("Variable 2")
    data3  = st.text_input("Variable 3")
    data4  = st.text_input("Variable 4")
    data5  = st.text_input("Variable 5")
    data6  = st.text_input("Variable 6")
    data7  = st.text_input("Variable 7")
    data8  = st.text_input("Variable 8")
    data9  = st.text_input("Variable 9")
    data10 = st.text_input("Variable 10")

    
    # Labels d'orientation
    st.write("\n---")
    st.write("Instructions :")
    st.write("- Assurez-vous de remplir tous les champs.")
    st.write("- Cliquez sur le bouton 'Envoyer' pour soumettre le formulaire.")
    
    # Bouton Envoyer
    if st.button("Envoyer"):
        # Traitement des données : enrégistrement dans la base de données
        process_data(data1, data2, data3, data4, data5, data6, data7, data8, data9, data10)
        st.write("\n---")
        st.write("Les données ont été bien enrégistrées\n")
        st.write("Merci pour votre contribution !")


if __name__ == '__main__':
    main()





